"""Lecture des fichiers sources I3F du pack AVP (Tarare 0546L).

Lit les .xlsx transmis par I3F et les normalise en structures mémoire
(tables à plat + scalaires de synthèse), **sans rien inventer** : un
fichier, un onglet, une colonne ou une valeur absent reste ``None`` et
sera rendu « Information non disponible dans les documents fournis. » par
les builders.

Design :

- ``_read_table`` détecte la ligne d'en-tête via une **ancre** (nom de
  colonne connu), préserve l'**ordre** des colonnes I3F, lit les lignes
  jusqu'à la première ligne totalement vide, et **écarte** les lignes de
  synthèse/notes (``Nombre de types…``, ``SHAB :``, ``= appuis…``).
- ``_scan_value`` récupère un scalaire de synthèse en cherchant un
  libellé (sous-chaîne) et en prenant la 1re valeur numérique à droite.

Aucune écriture ; ``read_only``.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from bim_core.contracts import load_envelope_quantities

from ..requirements._openpyxl_compat import patch_openpyxl

patch_openpyxl()

# Jetons signalant une ligne de synthèse / note à écarter d'une table.
_SUMMARY_TOKENS = (
    "nombre de types",
    "superficie des",
    "shab :",
    "shab:",
    "ratio ",
    "seuil ",
    "écart :",
    "ecart :",
    "= appuis",
    "somme de",
)


@dataclass
class SheetTable:
    """Table à plat : en-têtes ordonnés + lignes alignées."""

    title: str
    headers: list[str]
    rows: list[list[Any]] = field(default_factory=list)

    @property
    def n_rows(self) -> int:
        return len(self.rows)


@dataclass
class ControleMaquettesSource:
    template_path: Path | None = None
    header: dict[str, Any] = field(default_factory=dict)  # projet, esi, phase, dates, version
    grille: SheetTable | None = None
    legend: dict[int, str] = field(default_factory=dict)  # 0/1/2 -> libellé
    stats: dict[str, dict[str, Any]] = field(default_factory=dict)  # onglet -> stats conformité
    # Grille détaillée complète de chaque onglet de contrôle (listes
    # exploitables : noms de zones/pièces, éléments sans matériau…).
    stat_grids: dict[str, SheetGrid] = field(default_factory=dict)


@dataclass
class EnveloppeSource:
    table: SheetTable | None = None
    sheet_title: str | None = None  # nom d'onglet source (proximité I3F)
    superficie_facades: float | None = None
    superficie_menuiseries: float | None = None
    shab: float | None = None
    ratio_fac_shab: float | None = None
    seuil_3f: float | None = None
    # Enrichissements « logique MOA IfcOpenShell » (source structurée envelope.json).
    superficie_facades_nette: float | None = None
    superficie_calque_total: float | None = None  # total brut (hors filtre inclus)
    superficie_fenetres: float | None = None
    superficie_portes: float | None = None
    hors_filtre_type: list[dict] | None = None  # diagnostic, hors total métier
    # Périmètre de comptage des menuiseries (mode Revit : « murs extérieurs
    # avant filtre de type ») et part de ces menuiseries portée par un type
    # écarté. Sans ces deux valeurs, la colonne F du livrable peut être à zéro
    # sur toutes les lignes face à un total non nul, et se lire comme un bug.
    menuiseries_perimetre: str | None = None
    menuiseries_sur_types_rejetes: float | None = None
    # Filtre réellement appliqué par le backend (``diagnostics.filters``). La
    # note méthodologique du livrable en découle : elle décrit ce qui a été fait,
    # pas ce qu'on suppose avoir été fait.
    filter_mode: str | None = None
    filter_type_pattern: str | None = None
    filter_types_retenus: list[str] | None = None
    # Nature du DÉNOMINATEUR du ratio FAC/SHAB (``summary.methode_shab``). La
    # formule est unique, mais un ratio ne se compare qu'à un ratio dont la SHAB
    # est de même nature — d'où sa présence dans la note de méthode.
    methode_shab: str | None = None


@dataclass
class MenuiseriesSource:
    table: SheetTable | None = None
    sheet_title: str | None = None  # nom d'onglet source (proximité I3F)
    nombre_types: int | None = None
    # Notes de méthode écrites SOUS la table. Le gabarit A:N n'a qu'un onglet :
    # une explication ne peut pas y prendre la forme d'un onglet dédié. Sert
    # notamment à dire qu'un calcul existe mais a été jugé non comparable —
    # sinon une colonne vide se lit comme une absence de calcul.
    notes: list[str] = field(default_factory=list)


@dataclass
class SheetGrid:
    """Grille brute d'un onglet (reproduction fidèle multi-onglets)."""

    title: str
    rows: list[list[Any]] = field(default_factory=list)


@dataclass
class MultiSheetSource:
    """Ensemble d'onglets d'un export I3F (pivots + détail préservés)."""

    grids: list[SheetGrid] = field(default_factory=list)


@dataclass
class AvpSources:
    controle: ControleMaquettesSource | None = None
    shab: MultiSheetSource | None = None
    zones_espaces: MultiSheetSource | None = None
    enveloppe: EnveloppeSource | None = None
    menuiseries: MenuiseriesSource | None = None
    # Le classeur MOA « plancher » a **deux onglets** (« … Dalles Ok », « Planchers »
    # avec totaux/formules) → multi-onglets préservés, comme SHAB/Zones.
    plancher: MultiSheetSource | None = None


# ── Helpers de lecture ────────────────────────────────────────────────────


def _is_summary_row(vals: list[Any]) -> bool:
    for v in vals:
        if isinstance(v, str):
            lv = v.strip().lower()
            for tok in _SUMMARY_TOKENS:
                if tok and lv.startswith(tok):
                    return True
    return False


def _find_sheet(wb, *name_fragments: str):
    """Retourne le 1er onglet dont le titre contient un des fragments."""
    for ws in wb.worksheets:
        low = ws.title.lower()
        if any(frag.lower() in low for frag in name_fragments):
            return ws
    return None


def _read_table(ws, anchor: str, *, max_header_scan: int = 40) -> SheetTable | None:
    """Lit une table à plat en détectant l'en-tête via ``anchor``."""
    if ws is None:
        return None
    header_row = None
    for r in range(1, min(ws.max_row, max_header_scan) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and anchor.lower() in v.lower():
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        return None

    cols: list[tuple[int, str]] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip():
            cols.append((c, v.strip()))
    if not cols:
        return None
    idxs = [c for c, _ in cols]
    headers = [h for _, h in cols]

    rows: list[list[Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in idxs]
        if all(v in (None, "") for v in vals):
            break
        if _is_summary_row(vals):
            continue
        rows.append(vals)
    return SheetTable(title=ws.title, headers=headers, rows=rows)


def _scan_value(ws, *label_fragments: str) -> float | int | None:
    """Cherche un libellé (sous-chaîne) et renvoie la 1re valeur numérique
    à droite sur la même ligne. ``None`` si introuvable."""
    if ws is None:
        return None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and any(f.lower() in v.lower() for f in label_fragments):
                for c2 in range(c + 1, ws.max_column + 1):
                    v2 = ws.cell(r, c2).value
                    if isinstance(v2, (int, float)) and not isinstance(v2, bool):
                        return v2
    return None


def _open(path: str | Path):
    return openpyxl.load_workbook(str(path), read_only=False, data_only=True)


# ── Lecteurs par livrable ──────────────────────────────────────────────────


def read_controle(path: str | Path) -> ControleMaquettesSource:
    wb = _open(path)
    src = ControleMaquettesSource(template_path=Path(path))
    ws = _find_sheet(wb, "grille")
    if ws is not None:
        # Bloc entête (label en col B, valeur col C).
        header: dict[str, Any] = {}
        for r in range(1, min(ws.max_row, 12) + 1):
            label = ws.cell(r, 2).value
            val = ws.cell(r, 3).value
            if isinstance(label, str) and label.strip() and val not in (None, ""):
                header[label.strip().lower()] = val
        src.header = header
        # Légende 0/1/2 (col E numérique, col F libellé).
        for r in range(1, min(ws.max_row, 15) + 1):
            code = ws.cell(r, 5).value
            lib = ws.cell(r, 6).value
            if isinstance(code, int) and isinstance(lib, str) and lib.strip():
                src.legend[code] = lib.strip()
        src.grille = _read_table(ws, "POINTS DE CONTROLE")
    # Onglets de stats conformité.
    for sheet_name in (
        "Zones Nommage",
        "Pièces Nommage",
        "ARC bsence de matériau",
        "Zones ObjectType",
    ):
        ws_s = _find_sheet(wb, sheet_name.lower()) or (
            wb[sheet_name] if sheet_name in wb.sheetnames else None
        )
        if ws_s is None:
            continue
        materiau = "mat" in sheet_name.lower()
        src.stats[sheet_name] = _read_stats(ws_s, materiau=materiau)
        # Grille détaillée complète (listes de contrôle exploitables I3F).
        grid = _grid(ws_s)
        if grid:
            src.stat_grids[sheet_name] = SheetGrid(title=ws_s.title, rows=grid)
    wb.close()
    return src


def _read_stats(ws, *, materiau: bool) -> dict[str, Any]:
    """Extrait les stats conformité d'un onglet.

    Deux structures I3F distinctes :

    - **Nommage** (Zones/Pièces Nommage, Zones ObjectType) : ligne
      ``label | total | conforme | taux | non_conforme | taux`` (le label
      varie : « Noms (nbre) », « Nombre de Noms »…). On détecte la 1re
      ligne où col B est un libellé et col C/D/E sont numériques.
    - **Matériau** (« ARC bsence de matériau ») : ``label sans mat. | count
      | ratio`` + total à droite de « Nombre d'élements : ».
    """
    if materiau:
        out: dict[str, Any] = {}
        for r in range(1, min(ws.max_row, 12) + 1):
            lab = ws.cell(r, 2).value
            if isinstance(lab, str) and "sans mat" in lab.lower():
                out["label"] = lab.strip()
                out["non_conforme"] = ws.cell(r, 3).value
                out["non_conforme_ratio"] = ws.cell(r, 4).value
                break
        total = _scan_value(ws, "ements :")  # « Nombre d'élements : » (typo source)
        if total is not None:
            out["total"] = total
        return out if out.get("non_conforme") is not None else {}

    for r in range(1, min(ws.max_row, 14) + 1):
        lab = ws.cell(r, 2).value
        c3, c4, c5 = ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value
        num = lambda v: isinstance(v, (int, float)) and not isinstance(v, bool)  # noqa: E731
        if isinstance(lab, str) and lab.strip() and num(c3) and num(c4) and num(c5):
            return {
                "label": lab.strip(),
                "total": c3,
                "conforme": c4,
                "conforme_ratio": c5,
                "non_conforme": ws.cell(r, 6).value,
                "non_conforme_ratio": ws.cell(r, 7).value,
            }
    return {}


def _grid(ws) -> list[list[Any]]:
    """Grille brute (used range), triée des lignes/colonnes vides finales."""
    rows: list[list[Any]] = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)] for r in range(1, ws.max_row + 1)
    ]
    while rows and all(v in (None, "") for v in rows[-1]):
        rows.pop()
    if not rows:
        return []
    last_col = 0
    for row in rows:
        for i, v in enumerate(row):
            if v not in (None, ""):
                last_col = max(last_col, i + 1)
    return [row[:last_col] for row in rows]


def read_all_grids(path: str | Path) -> list[SheetGrid]:
    """Lit **tous** les onglets d'un classeur en grilles brutes.

    Préserve **tous** les onglets I3F, y compris vides (structure Excel
    stricte : pivots ``Feuil1``/``Feuil2`` + détail ``TDB``). Un onglet
    vide donne un ``SheetGrid`` à lignes vides.
    """
    wb = _open(path)
    out: list[SheetGrid] = [SheetGrid(title=ws.title, rows=_grid(ws)) for ws in wb.worksheets]
    wb.close()
    return out


def read_shab(path: str | Path) -> MultiSheetSource:
    return MultiSheetSource(grids=read_all_grids(path))


def read_zones_espaces(path: str | Path) -> MultiSheetSource:
    return MultiSheetSource(grids=read_all_grids(path))


def read_enveloppe(path: str | Path) -> EnveloppeSource:
    wb = _open(path)
    ws = wb.worksheets[0]
    src = EnveloppeSource(table=_read_table(ws, "Composant"), sheet_title=ws.title)
    src.superficie_facades = _scan_value(ws, "superficie des façades")
    src.superficie_menuiseries = _scan_value(ws, "superficie des menuiseries")
    src.shab = _scan_value(ws, "shab")
    src.ratio_fac_shab = _scan_value(ws, "ratio fac/shab")
    src.seuil_3f = _scan_value(ws, "seuil 3f")
    wb.close()
    return src


# Onglet + colonnes MOA « Extraction surface enveloppe » (logique Tarare, sans
# Solibri : les colonnes Solibri deviennent des colonnes IFC OpenShell).
ENVELOPPE_MOA_SHEET = "TDB 2022 04.2 - Extraction s..."
ENVELOPPE_MOA_HEADERS = [
    "Composant",  # A
    "Type",  # B
    "Étages",  # C
    "Archicad BQ NetSideArea",  # D
    "Surface IFC OpenShell",  # E (ex-« Surface Solibri »)
    "ArchiCAD Superficie des ouvertures sur face extérieure",  # F
    "IFC OpenShell Surface des Fenêtres",  # G (ex-« Solibri … »)
    "IFC OpenShell Surface des Portes",  # H (ex-« Solibri … »)
    "Nombre",  # I
    "Couleur",  # J
]


def read_envelope_json(path: str | Path) -> EnveloppeSource:
    """Construit la source enveloppe MOA depuis le contrat ``envelope_quantities/v1``.

    **Une ligne métier par type** (``par_type``), jamais par mur élémentaire ;
    ``hors_filtre_type`` reste hors du total métier (diagnostic).

    La lecture est déléguée à :func:`bim_core.contracts.load_envelope_quantities`,
    qui porte la politique de schéma commune : document V1 accepté, schéma
    inconnu **refusé**, fichier historique sans ``schema`` migré vers V1 avec
    l'avertissement ``legacy_schema_missing``. La normalisation des alias de
    clés (``netsidearea_m2``, ``nombre``, ``etages`` en chaîne…) vit désormais
    dans ce contrat — elle n'est plus dupliquée ici.

    Raises:
        ContractError: schéma inconnu/invalide, forme non reconnue, ou fichier
            illisible (sous-classe de ``ValueError``).
    """
    payload = load_envelope_quantities(str(path))
    summary = payload.summary

    rows: list[list[Any]] = []
    for e in sorted(payload.par_type, key=lambda r: str(r.type or "")):
        rows.append(
            [
                "Mur",  # A Composant (libellé MOA)
                e.type,  # B Type
                ", ".join(e.etages),  # C Étages
                e.net_side_area_m2,  # D Archicad BQ NetSideArea
                # E : aire recalculée si le producteur la ventile, sinon la
                # valeur Archicad — la colonne ne reste pas vide pour autant.
                (
                    e.surface_ifc_openshell_m2
                    if e.surface_ifc_openshell_m2 is not None
                    else e.net_side_area_m2
                ),
                # F : ouvertures extérieures si ventilées par type, sinon le
                # total menuiseries du type.
                (
                    e.superficie_ouvertures_exterieures_m2
                    if e.superficie_ouvertures_exterieures_m2 is not None
                    else e.menuiseries_m2
                ),
                e.fenetres_m2,  # G fenêtres
                e.portes_m2,  # H portes
                e.n,  # I Nombre
                None,  # J Couleur
            ]
        )

    table = SheetTable(title=ENVELOPPE_MOA_SHEET, headers=list(ENVELOPPE_MOA_HEADERS), rows=rows)
    return EnveloppeSource(
        table=table,
        sheet_title=ENVELOPPE_MOA_SHEET,
        superficie_facades=summary.superficie_facades_m2,
        superficie_menuiseries=summary.superficie_menuiseries_m2,
        shab=summary.shab_m2,
        ratio_fac_shab=summary.ratio_fac_shab,
        seuil_3f=summary.seuil_i3f,
        superficie_facades_nette=summary.superficie_facades_nette_m2,
        superficie_calque_total=summary.superficie_calque_total_m2,
        superficie_fenetres=summary.superficie_menuiseries_fenetres_m2,
        superficie_portes=summary.superficie_menuiseries_portes_m2,
        hors_filtre_type=[r.model_dump() for r in payload.hors_filtre_type],
        # ``methode_shab`` est un champ ``extra`` du summary V1 : les contrats
        # antérieurs à ifc-geometry-mcp v0.5.0 ne le portent pas, et l'absence
        # doit rester silencieuse plutôt que d'affirmer une méthode inconnue.
        methode_shab=_texte_ou_none(getattr(summary, "methode_shab", None)),
        **_menuiseries_perimetre(payload),
    )


def _menuiseries_perimetre(payload) -> dict[str, Any]:
    """Périmètre de comptage des menuiseries, si le producteur le déclare.

    Émis par le mode ``geometric_type_filter`` : les baies y sont comptées sur
    les murs extérieurs **avant** filtre de type, parce qu'en façade Revit
    multicouche elles sont portées par le mur porteur et non par la peau. Sans
    cette information, le livrable afficherait une colonne « ouvertures » nulle
    sur toutes les lignes face à un total non nul.
    """
    diagnostics = getattr(payload, "diagnostics", None) or {}
    if not isinstance(diagnostics, dict):
        return {}
    perimetre = diagnostics.get("menuiseries_perimetre")
    rejetes = diagnostics.get("menuiseries_m2_sur_types_rejetes")
    filtres = diagnostics.get("filters") or {}
    if not isinstance(filtres, dict):
        filtres = {}
    retenus = filtres.get("types_retenus")
    return {
        "menuiseries_perimetre": perimetre if isinstance(perimetre, str) else None,
        "menuiseries_sur_types_rejetes": (
            float(rejetes) if isinstance(rejetes, (int, float)) else None
        ),
        "filter_mode": _texte_ou_none(filtres.get("mode")),
        "filter_type_pattern": _texte_ou_none(filtres.get("type_pattern")),
        "filter_types_retenus": ([str(t) for t in retenus] if isinstance(retenus, list) else None),
    }


def _texte_ou_none(valeur) -> str | None:
    return valeur.strip() if isinstance(valeur, str) and valeur.strip() else None


def read_menuiseries(path: str | Path) -> MenuiseriesSource:
    wb = _open(path)
    ws = wb.worksheets[0]
    src = MenuiseriesSource(table=_read_table(ws, "Composant"), sheet_title=ws.title)
    nb = _scan_value(ws, "nombre de types")
    src.nombre_types = int(nb) if isinstance(nb, (int, float)) else None
    wb.close()
    return src


def read_plancher(path: str | Path) -> MultiSheetSource:
    """Lit l'export plancher I3F en **préservant tous ses onglets**.

    Le classeur MOA « plancher » porte deux onglets (« … Dalles Ok » et
    « Planchers » avec totaux/écarts d'outil externe). Comme SHAB/Zones, on conserve la
    structure multi-onglets plutôt que d'aplatir sur le seul premier onglet.
    """
    return MultiSheetSource(grids=read_all_grids(path))


@dataclass
class AvpSourcePaths:
    """Chemins des 5 .xlsx sources I3F (tous optionnels)."""

    controle: str | Path | None = None
    shab: str | Path | None = None
    zones_espaces: str | Path | None = None
    enveloppe: str | Path | None = None
    menuiseries: str | Path | None = None
    plancher: str | Path | None = None


def load_sources(paths: AvpSourcePaths) -> AvpSources:
    """Charge les sources disponibles ; chaque livrable absent reste ``None``."""
    out = AvpSources()
    if paths.controle:
        out.controle = read_controle(paths.controle)
    if paths.shab:
        out.shab = read_shab(paths.shab)
    if paths.zones_espaces:
        out.zones_espaces = read_zones_espaces(paths.zones_espaces)
    if paths.enveloppe:
        out.enveloppe = read_enveloppe(paths.enveloppe)
    if paths.menuiseries:
        out.menuiseries = read_menuiseries(paths.menuiseries)
    if paths.plancher:
        out.plancher = read_plancher(paths.plancher)
    return out

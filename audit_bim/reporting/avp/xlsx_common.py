"""Helpers xlsx bas-niveau et utilitaires texte/format partagés (charte MOA).

Briques réutilisées par 2 builders ou plus : écriture de tables/grilles MOA,
formats xlsxwriter, normalisation de texte, comptage QA. Dépend uniquement de
:mod:`models` et de l'infra reporting.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import xlsxwriter

from ...requirements._openpyxl_compat import patch_openpyxl
from ..avp_snapshot import _SRC_COMPUTED
from ..avp_sources import SheetTable
from ..word_report import NOT_AVAILABLE
from ..xlsx_annex import _build_formats, write_safe
from .models import _COMPUTED_METHODO_NOTE, _QA_SCAFFOLD

# Les annexes AVP fournies par la MOA peuvent porter les mêmes CustomFilter
# invalides que les fichiers I3F — patch appliqué explicitement (plus d'effet
# de bord d'import depuis que les parseurs requirements sont paresseux).
patch_openpyxl()


#: Motif des en-têtes de colonnes portant une quantité **calculée**.
_COL_OPENSHELL_TOKEN = "IFC OpenShell"
#: En-tête de l'ancien mécanisme de traçabilité (provenance écrite en toutes
#: lettres en bout de ligne). Sa présence identifie une table **non migrée**.
_COL_SOURCE_QUANTITE = "Source quantité"


def _rows_have_computed(rows, headers=None) -> bool:
    """Vrai si une table porte au moins une quantité calculée par IfcOpenShell.

    Deux mécanismes de traçabilité coexistent, et la table dit lequel elle
    emploie :

    - **non migrée** — une colonne ``Source quantité`` porte la provenance en
      toutes lettres. C'est encore le cas du livrable Plancher, dont les deux
      colonnes de mesure reçoivent la même valeur : y lire un calcul dans la
      colonne « IFC OpenShell » qualifierait de calculée une quantité native ;
    - **migrée** (doctrine #210) — la provenance se lit à **l'emplacement de la
      valeur**, une seule des deux colonnes étant renseignée. C'est le cas des
      livrables Zones/Espaces, SHAB et Fenêtres.

    Ne chercher que l'ancien libellé faisait disparaître la note « valeurs NON
    contractuelles » des livrables migrés — sans erreur, la note étant
    facultative.

    ``headers`` est explicite car les deux porteurs diffèrent : ``SheetGrid``
    inclut sa ligne d'en-tête dans ``rows``, ``SheetTable`` la garde à part.
    """
    rows = rows or []
    entetes = [e if isinstance(e, str) else "" for e in (headers if headers is not None else [])]
    donnees = rows
    if headers is None:
        entetes = [e if isinstance(e, str) else "" for e in (rows[0] if rows else [])]
        donnees = rows[1:]

    if any(e == _COL_SOURCE_QUANTITE for e in entetes):
        return any(c == _SRC_COMPUTED for r in rows for c in (r or []))

    colonnes = [i for i, entete in enumerate(entetes) if _COL_OPENSHELL_TOKEN in entete]
    return any(
        i < len(r or []) and isinstance(r[i], (int, float)) and not isinstance(r[i], bool)
        for r in donnees
        for i in colonnes
    )


def _count_business_rows(path: Path) -> int:
    """Ouvre une annexe et compte ses **lignes métier**.

    Ignore le bandeau éventuel, la ligne d'en-tête de chaque onglet, les
    marqueurs d'échafaudage (``NOT_AVAILABLE``, onglet vide) et les blocs KPI.
    Sert de garde qualité anti-livrable vide.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return 0
    total = 0
    try:
        for ws in wb.worksheets:
            header_seen = False
            for row in ws.iter_rows(values_only=True):
                cells = [c for c in row if c not in (None, "")]
                if not cells:
                    continue
                first = str(cells[0]).strip().lower()
                if first == "synthèse":
                    break  # début du bloc KPI → stop pour cet onglet
                if not header_seen and any(str(c).strip().lower() == "composant" for c in cells):
                    header_seen = True
                    continue
                if not header_seen:
                    continue
                if first in _QA_SCAFFOLD:
                    continue
                if row[0] in (None, ""):
                    continue
                total += 1
    finally:
        wb.close()
    return total


def _cell(v):
    """Valeur cellule sûre : blanc pour vide, date ISO, sinon brut."""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def _write_banner(ws, fmts, supertitle: str, title: str) -> int:
    """Bannière BIMData (supertitle + filet jaune + titre). Renvoie la
    prochaine ligne libre."""
    write_safe(ws, 0, 0, f"BIMDATA — {supertitle}", fmts["supertitle"])
    write_safe(ws, 1, 0, "", fmts["accent_filet"])
    ws.set_row(1, 4)  # filet jaune fin
    write_safe(ws, 2, 0, title, fmts["title"])
    return 4


def _write_flat_table(ws, fmts, table: SheetTable | None, *, start_row: int) -> int:
    """Écrit une table à plat (en-têtes brandés + lignes zébrées).

    ``table is None`` → mention ``NOT_AVAILABLE``. Renvoie la ligne suivante.
    """
    if table is None or not table.headers:
        write_safe(ws, start_row, 0, NOT_AVAILABLE, fmts["row"])
        return start_row + 1
    for c, h in enumerate(table.headers):
        write_safe(ws, start_row, c, h, fmts["header"])
        ws.set_column(c, c, max(12, min(42, len(str(h)) + 3)))
    ws.set_row(start_row, 28)
    r = start_row
    for i, rowvals in enumerate(table.rows):
        r = start_row + 1 + i
        fmt = fmts["row_alt"] if i % 2 == 0 else fmts["row"]
        for c, v in enumerate(rowvals):
            write_safe(ws, r, c, _cell(v), fmt)
    ws.freeze_panes(start_row + 1, 0)
    return r + 1


def _new_workbook(path: Path):
    wb = xlsxwriter.Workbook(str(path), {"strings_to_formulas": False})
    return wb, _build_formats(wb)


def _openpyxl_safe_value(value):
    value = _cell(_moa_text(value))
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def _text_or_none(value) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def _dict_text(item: dict, *keys: str) -> str | None:
    for key in keys:
        value = item.get(key)
        text = _text_or_none(value)
        if text:
            return text
    return None


def _write_stats_block(ws, fmts, stats: dict | None, *, start_row: int) -> None:
    if not stats:
        write_safe(ws, start_row, 0, NOT_AVAILABLE, fmts["row"])
        return
    # Structure « nommage » (avec conforme) ou « matériau » (sans conforme).
    if "conforme" in stats:
        labels = [
            ("Indicateur", "label"),
            ("Total", "total"),
            ("Conforme", "conforme"),
            ("Taux conforme", "conforme_ratio"),
            ("Non conforme", "non_conforme"),
            ("Taux non conforme", "non_conforme_ratio"),
        ]
    else:
        labels = [
            ("Indicateur", "label"),
            ("Total éléments", "total"),
            ("Sans matériau", "non_conforme"),
            ("Taux sans matériau", "non_conforme_ratio"),
        ]
    for c, (title, _key) in enumerate(labels):
        write_safe(ws, start_row, c, title, fmts["header"])
        ws.set_column(c, c, 20)
    for c, (_title, key) in enumerate(labels):
        v = stats.get(key)
        write_safe(ws, start_row + 1, c, "" if v is None else v, fmts["row_alt"])


def _looks_like_header(vals: list) -> bool:
    return sum(1 for v in vals if isinstance(v, str) and v.strip()) >= 3


def _write_grid(ws, fmts, rows: list[list], *, start_row: int) -> int:
    """Reproduit une grille brute (pivot/synthèse I3F) en table à plat.

    La 1re ligne « en-tête » (≥ 3 cellules texte) est stylée ; les autres
    sont zébrées. Préserve l'ordre et le contenu source.
    """
    if not rows:
        write_safe(ws, start_row, 0, NOT_AVAILABLE, fmts["row"])
        return start_row + 1
    header_idx = next((i for i, r in enumerate(rows) if _looks_like_header(r)), None)
    ncols = max(len(r) for r in rows)
    for c in range(ncols):
        ws.set_column(c, c, 18)
    r = start_row
    for i, rowvals in enumerate(rows):
        r = start_row + i
        if i == header_idx:
            fmt = fmts["header"]
            ws.set_row(r, 26)
        else:
            fmt = fmts["row_alt"] if i % 2 == 0 else fmts["row"]
        for c in range(ncols):
            v = rowvals[c] if c < len(rowvals) else None
            write_safe(ws, r, c, _cell(v), fmt)
    if header_idx is not None:
        ws.freeze_panes(start_row + header_idx + 1, 0)
    return r + 1


def _moa_formats(wb) -> dict[str, xlsxwriter.format.Format]:
    base = {"font_name": "Aptos Narrow", "font_size": 11}
    return {
        "title": wb.add_format({**base, "bold": True, "font_size": 13}),
        "note": wb.add_format({**base, "text_wrap": True}),
        "meta_label": wb.add_format({**base, "bold": True}),
        "meta_value": wb.add_format({**base}),
        "header": wb.add_format(
            {
                **base,
                "bold": True,
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
            }
        ),
        "data": wb.add_format(
            {**base, "border": 1, "valign": "top", "text_wrap": True, "num_format": "#,##0.00"}
        ),
        "center": wb.add_format({**base, "border": 1, "align": "center", "valign": "vcenter"}),
        "percent": wb.add_format({**base, "border": 1, "num_format": "0.00%"}),
    }


_MOA_FORMULA_PREFIXES = ("=IF(", "=SUBTOTAL(", "=COUNTA(", "=SUM(")


def _moa_text(value):
    if not isinstance(value, str):
        return value
    return (
        value.replace("Surface (Solibri)", "Surface IFC OpenShell")
        .replace("Surface\nSolibri", "Surface IFC OpenShell")
        .replace("Surface Solibri", "Surface IFC OpenShell")
        .replace("Solibri Surface des Fenêtres", "IFC OpenShell Surface des Fenêtres")
        .replace("Solibri Surface des Portes", "IFC OpenShell Surface des Portes")
    )


def _is_moa_formula(value) -> bool:
    return isinstance(value, str) and value.startswith(_MOA_FORMULA_PREFIXES)


def _write_moa_value(ws, row: int, col: int, value, fmt) -> None:
    if _is_moa_formula(value):
        ws.write_formula(row, col, value, fmt, "")
    else:
        write_safe(ws, row, col, _cell(_moa_text(value)), fmt)


def _write_moa_grid(ws, fmts, rows: list[list], *, start_row: int = 0) -> int:
    """Écrit une grille façon MOA depuis A1, sans bannière ni freeze panes."""
    if not rows:
        return start_row
    ncols = max(len(r) for r in rows)
    header_idx = next((i for i, r in enumerate(rows) if _looks_like_header(r)), 0)
    for c in range(ncols):
        sample = [r[c] for r in rows[:30] if c < len(r) and r[c] not in (None, "")]
        width = max([len(str(_moa_text(v))) for v in sample] + [10])
        ws.set_column(c, c, min(max(width + 2, 10), 42))
    for i, rowvals in enumerate(rows):
        xls_row = start_row + i
        fmt = fmts["header"] if i == header_idx else fmts["data"]
        if i == header_idx:
            ws.set_row(xls_row, 42 if i == 0 else 28)
        for c in range(ncols):
            value = rowvals[c] if c < len(rowvals) else None
            cell_fmt = fmts["percent"] if _is_moa_percent_formula(value) else fmt
            _write_moa_value(ws, xls_row, c, value, cell_fmt)
    return start_row + len(rows)


def _is_moa_percent_formula(value) -> bool:
    return isinstance(value, str) and "/D" in value and value.startswith("=IF(")


def _build_multisheet_export_xlsx(path, banner: str, title: str, multi, meta) -> Path:
    """Export reproduisant **tous** les onglets source (pivots + détail)."""
    wb = xlsxwriter.Workbook(str(path), {"strings_to_formulas": False})
    fmts = _moa_formats(wb)
    grids = (multi.grids if multi else None) or []
    if not grids:
        ws = wb.add_worksheet(_safe_sheet(title))
        write_safe(ws, 0, 0, NOT_AVAILABLE, fmts["data"])
        wb.close()
        return path
    for g in grids:
        ws = wb.add_worksheet(_safe_sheet(g.title))
        if g.rows:
            end = _write_moa_grid(ws, fmts, g.rows, start_row=0)
            if _rows_have_computed(g.rows):
                write_safe(ws, end + 1, 0, _COMPUTED_METHODO_NOTE, fmts["note"])
    wb.close()
    return path


def _first_number(*values) -> float | None:
    for value in values:
        if isinstance(value, (int, float)):
            return float(value)
    return None


def _formula_cached(value):
    return value if isinstance(value, (int, float)) else ""


def _sum_table_col(rows, idx: int) -> float:
    return round(sum(r[idx] for r in rows if len(r) > idx and isinstance(r[idx], (int, float))), 2)


def _pct(v) -> str:
    return f"{v * 100:.0f} %" if isinstance(v, (int, float)) else NOT_AVAILABLE


def _fmt_meta(v) -> str:
    if v in (None, ""):
        return NOT_AVAILABLE
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v)


def _norm(s: str) -> str:
    return "".join(ch for ch in s.lower() if ch.isalnum())


def _safe_sheet(title: str) -> str:
    bad = set(r"[]:*?/\\")
    return "".join(c for c in title if c not in bad)[:31] or "Feuille"


def _stat_lookup(ctrl, name: str) -> dict:
    if not ctrl or not ctrl.stats:
        return {}
    for key, val in ctrl.stats.items():
        if _norm(key).replace("bsence", "absence") == _norm(name):
            return val or {}
    return {}

"""Un pack dont les colonnes de quantités sont vides n'est pas livrable.

Le cas est plus dangereux qu'un pack vide : les annexes ont leurs lignes, leurs
en-têtes et leur mise en forme — elles se lisent comme un résultat, alors que
« Surface IFC OpenShell », « Surface Nette (Qté de Base) »,
« BaseQuantities.Width/Height » et « BaseQuantities.NetArea » ne contiennent
rien. La QA gate historique ne voyait pas ce cas : elle ne comptait que les
lignes.

La cause est en amont : le snapshot BIMData ne porte pas de ``BaseQuantities``
et les quantités calculées (contrat ``computed_base_quantities/v1``) n'ont pas
été fusionnées.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.mcp.session import _Session, current_session
from audit_bim.profiles.i3f.tools_reporting import generate_avp_i3f_pack as tr_generate_avp_i3f_pack
from audit_bim.profiles.i3f.tools_reporting import (
    list_avp_i3f_xls_reports as tr_list_avp_i3f_xls_reports,
)

SOURCE_CALCULEE = "Calculée (IfcOpenShell)"


@pytest.fixture
def session(tmp_path, monkeypatch):
    monkeypatch.setenv("AUDIT_OUTPUT_DIR", str(tmp_path))
    monkeypatch.delenv("AUDIT_INPUT_DIR", raising=False)
    sess = _Session()
    token = current_session.set(sess)
    try:
        yield sess, tmp_path
    finally:
        current_session.reset(token)


def _snapshot_sans_quantites(*, pset_natif_sp1: list[dict] | None = None) -> ModelSnapshot:
    """Snapshot BIMData réaliste : entités présentes, AUCUNE BaseQuantity.

    ``pset_natif_sp1`` permet de doter SP1 d'une quantité **native** pour
    vérifier que la fusion reste gap-only.
    """
    sp1 = {"uuid": "SP1", "type": "IfcSpace", "name": "SEJOUR", "longname": "SEJOUR"}
    if pset_natif_sp1:
        sp1["property_sets"] = pset_natif_sp1
    spaces = [
        sp1,
        {"uuid": "SP2", "type": "IfcSpace", "name": "CHAMBRE 01", "longname": "CHAMBRE 01"},
    ]
    elements = [
        {"uuid": "W1", "type": "IfcWindow", "name": "F25"},
        {"uuid": "D1", "type": "IfcDoor", "name": "P10"},
        {"uuid": "SL1", "type": "IfcSlab", "name": "Dalle RDC"},
    ]
    return ModelSnapshot(
        project={"name": "Dieppe"},
        model={"name": "DIEPPE-7427L.ifc"},
        spaces=spaces,
        elements=elements,
    ).index()


def _computed_json(path: Path) -> str:
    """Contrat ``computed_base_quantities/v1`` couvrant les 4 annexes."""
    doc = {
        "schema": "computed_base_quantities/v1",
        "source": {
            "producer": "ifc-geometry",
            "tool": "export_computed_base_quantities",
            "version": "0.2.0",
            "ifc_file": "DIEPPE-7427L.ifc",
        },
        "created_at": "2026-08-01T20:00:00+00:00",
        "quantities": [
            _q("SP1", "IfcSpace", "Qto_SpaceBaseQuantities", "NetFloorArea", 24.5),
            _q("SP2", "IfcSpace", "Qto_SpaceBaseQuantities", "NetFloorArea", 12.98),
            _q("W1", "IfcWindow", "Qto_WindowBaseQuantities", "Width", 0.6),
            _q("W1", "IfcWindow", "Qto_WindowBaseQuantities", "Height", 1.3),
            _q("D1", "IfcDoor", "Qto_DoorBaseQuantities", "Width", 0.9),
            _q("D1", "IfcDoor", "Qto_DoorBaseQuantities", "Height", 2.1),
            _q("SL1", "IfcSlab", "Qto_SlabBaseQuantities", "NetArea", 156.4),
        ],
        "coverage": {"n_elements": 5, "n_computed": 7, "n_failed": 0},
        "warnings": [],
    }
    path.write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")
    return str(path)


def _q(gid, cls, qto, name, value):
    return {
        "global_id": gid,
        "ifc_class": cls,
        "qto": qto,
        "quantity": name,
        "value": value,
        "unit": "m2" if "Area" in name else "m",
        "method": "geometry",
        "status": "computed",
        "source": "computed_ifcopenshell",
    }


def _cells(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    valeurs = [c for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for c in row]
    wb.close()
    return valeurs


def _nombres(path) -> list[float]:
    return [c for c in _cells(path) if isinstance(c, (int, float)) and not isinstance(c, bool)]


def _generer(*, computed=None, auto=False, **kw):
    """Génération de test. ``auto=False`` par défaut : ces tests portent sur la
    QA gate elle-même, pas sur l'auto-résolution (fichier dédié)."""
    return tr_generate_avp_i3f_pack(
        auto_compute_quantities=auto,
        auto_compute_envelope=auto,
        project_name="Dieppe",
        project_code="7427L",
        phase="APD",
        auditor_name="Stanislas Limouzi",
        export_pdf=False,
        computed_quantities_json=computed,
        **kw,
    )


# ── sans quantités : refus explicite, pas un pack faux ─────────────────


def test_pack_without_quantities_is_refused(session):
    sess, tmp_path = session
    sess.snapshot = _snapshot_sans_quantites()

    res = _generer()

    assert res["status"] == "error"
    assert res["error"] == "missing_quantities"
    assert res["needs_computed_quantities_json"] is True
    # Les quatre annexes concernées sont nommées.
    assert set(res["empty_deliverables"]) == {"SHAB", "Zones/Espaces", "Menuiseries", "Plancher"}
    assert "computed_quantities_json" in res["next_step"]


def test_refusal_message_names_the_producing_tool(session):
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    res = _generer()

    assert "export_computed_base_quantities" in res["message"]


def test_availability_tool_signals_the_need_upfront(session):
    """Le besoin remonte AVANT la génération, pas au moment du refus."""
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    res = tr_list_avp_i3f_xls_reports()

    assert res["needs_computed_quantities_json"] is True
    assert set(res["reports_without_quantities"]) == {
        "SHAB",
        "Zones/Espaces",
        "Menuiseries",
        "Plancher",
    }
    assert "export_computed_base_quantities" in res["next_action"]


# ── avec le JSON calculé : colonnes réellement remplies ────────────────


def test_pack_with_computed_json_is_generated(session, tmp_path):
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    res = _generer(computed=_computed_json(tmp_path / "computed.json"))

    assert res.get("status") != "error", res
    assert res["computed_quantities_json_used"].endswith("computed.json")
    couverture = res["computed_quantities_coverage"]
    assert couverture["n_merged"] == 7
    assert couverture["n_unknown_uuid"] == 0


@pytest.mark.parametrize(
    ("annexe", "valeurs_attendues"),
    [
        ("shab_xlsx", (24.5, 12.98)),
        ("zones_espaces_xlsx", (24.5, 12.98)),
        # 0.9 x 2.1 était la PORTE : ce livrable a pour gabarit
        # « Fenêtres Ok » et ne collecte plus que des IfcWindow.
        ("menuiseries_xlsx", (0.6, 1.3)),
        ("plancher_xlsx", (156.4,)),
    ],
)
def test_quantity_columns_contain_real_values(session, tmp_path, annexe, valeurs_attendues):
    """Chaque annexe porte les valeurs fusionnées — pas seulement des lignes."""
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    res = _generer(computed=_computed_json(tmp_path / "computed.json"))
    assert res.get("status") != "error", res

    chemin = next(p for p in res["paths"] if _cle(p) == annexe)
    nombres = _nombres(chemin)
    for attendue in valeurs_attendues:
        assert any(abs(n - attendue) < 0.01 for n in nombres), (
            f"{attendue} absente de {Path(chemin).name} (valeurs : {sorted(set(nombres))[:12]})"
        )


def _cle(chemin: str) -> str:
    nom = Path(chemin).name
    for cle, libelle in (
        ("shab_xlsx", "export SHAB maquette"),
        ("zones_espaces_xlsx", "Export Zones et Espaces"),
        ("menuiseries_xlsx", "export Menuiseries"),
        ("plancher_xlsx", "export plancher"),
    ):
        if libelle in nom:
            return cle
    return ""


@pytest.mark.parametrize("annexe", ["shab_xlsx", "zones_espaces_xlsx"])
def test_quantity_source_is_traced_as_computed(session, tmp_path, annexe):
    """Une quantité calculée est tracée comme telle dans l'onglet de détail.

    La provenance ne se lit plus dans une colonne « Source quantité » en bout
    de ligne — absente du gabarit client, elle en déformait le tableau — mais à
    **l'emplacement de la valeur** : ``Surface IFC OpenShell`` pour le calculé,
    ``Surface Nette (Qté de Base)`` pour le natif, jamais les deux (doctrine
    #210). L'exigence est inchangée : ce qui est interdit, c'est qu'une valeur
    fusionnée sorte sans provenance — le symptôme d'origine (299 fois
    « Information non disponible » dans cette colonne).
    """
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    res = _generer(computed=_computed_json(tmp_path / "computed.json"))
    assert res.get("status") != "error", res

    chemin = next(p for p in res["paths"] if _cle(p) == annexe)
    wb = openpyxl.load_workbook(chemin)
    ws = next(wb[t] for t in wb.sheetnames if t.startswith("TDB 2022 01.3"))
    entetes = [c.value for c in ws[1]]
    col_calc = entetes.index("Surface IFC OpenShell") + 1
    col_natif = entetes.index("Surface Nette (Qté de Base)") + 1
    assert "Source quantité" not in entetes

    calculees = natives = 0
    for ligne in range(2, ws.max_row + 1):
        v_calc = ws.cell(ligne, col_calc).value
        v_natif = ws.cell(ligne, col_natif).value
        if v_calc is None and v_natif is None:
            continue  # ligne de séparation / sous-total
        assert (v_calc is None) != (v_natif is None), (
            f"ligne {ligne} de {Path(chemin).name} : les deux colonnes de "
            "mesure sont remplies, la provenance n'est plus lisible"
        )
        calculees += v_calc is not None
        natives += v_natif is not None

    assert calculees >= 2, (
        f"surfaces fusionnées attendues en « Surface IFC OpenShell » dans "
        f"{Path(chemin).name} (trouvé {calculees}, natives {natives})"
    )


# ── la fusion reste gap-only ───────────────────────────────────────────


def test_native_quantities_are_never_overwritten(session, tmp_path):
    """Une BaseQuantity native BIMData prime sur la valeur calculée.

    *Gap-only* décide quelle valeur fait **autorité**, pas laquelle mérite
    d'être retenue : la native reste dans le pset et dans sa colonne, et la
    calculée est désormais **conservée à côté** pour alimenter la colonne
    « IFC OpenShell ». C'est ce qui rend la comparaison possible — jusqu'ici la
    valeur calculée était jetée dès qu'une native existait, donc la colonne de
    comparaison sortait vide sur toute maquette portant ses BaseQuantities.
    """
    sess, _ = session
    # La quantité native vit dans les DONNÉES du snapshot (ce que renvoie
    # BIMData), pas dans l'index : ``index()`` recopie les espaces
    # (``{**it, "type": kind}``), donc muter ``element_by_uuid`` ne
    # modifierait qu'un artefact reconstruit à la copie.
    sess.snapshot = _snapshot_sans_quantites(
        pset_natif_sp1=[
            {
                "name": "Qto_SpaceBaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 99.9}],
            }
        ]
    )

    res = _generer(computed=_computed_json(tmp_path / "computed.json"))
    assert res.get("status") != "error", res
    assert res["computed_quantities_coverage"]["n_gap_kept"] == 1

    chemin = next(p for p in res["paths"] if _cle(p) == "shab_xlsx")
    wb = openpyxl.load_workbook(chemin)
    ws = next(wb[t] for t in wb.sheetnames if t.startswith("TDB 2022 01.3"))
    entetes = [c.value for c in ws[1]]
    col_natif = entetes.index("Surface Nette (Qté de Base)") + 1
    col_calc = entetes.index("Surface IFC OpenShell") + 1
    valeurs = [
        (ws.cell(r, col_natif).value, ws.cell(r, col_calc).value) for r in range(2, ws.max_row + 1)
    ]
    # La native reste à SA place, la calculée n'y entre pas.
    assert any(n is not None and abs(n - 99.9) < 0.01 for n, _ in valeurs), (
        "la valeur native doit être conservée dans la colonne native"
    )
    assert not any(n is not None and abs(n - 24.5) < 0.01 for n, _ in valeurs), (
        "la calculée ne doit pas écraser la native"
    )
    # …et elle est visible en face, ce qui rend l'écart comparable.
    assert any(c is not None and abs(c - 24.5) < 0.01 for _, c in valeurs), (
        "la valeur calculée doit alimenter la colonne « Surface IFC OpenShell »"
    )


def test_unknown_schema_is_refused_before_generating(session, tmp_path):
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()
    mauvais = tmp_path / "bad.json"
    mauvais.write_text(json.dumps({"schema": "autre/v9", "quantities": []}), encoding="utf-8")

    with pytest.raises(ValueError, match="Schéma non reconnu"):
        _generer(computed=str(mauvais))


# ── le refus ne laisse RIEN sur disque ─────────────────────────────────


def test_refusal_writes_no_deliverable_at_all(session, tmp_path):
    """Un refus doit être un non-événement sur le disque.

    Contrôler après génération renverrait bien ``status="error"``, mais un
    dossier de livrables non conformes existerait — celui-là même qu'on
    cherche à ne jamais produire. La gate est donc en **préflight**.
    """
    sess, out_root = session
    sess.snapshot = _snapshot_sans_quantites()
    cible = "pack_refuse"

    res = _generer(output_dir=cible)

    assert res["error"] == "missing_quantities"
    dossier = Path(out_root) / cible
    livrables = list(Path(out_root).rglob("*.xlsx")) + list(Path(out_root).rglob("*.docx"))
    assert livrables == [], f"livrables écrits malgré le refus : {livrables}"
    assert not dossier.exists(), "le dossier de sortie ne doit pas être créé"


# ── rejouabilité : un second JSON doit primer ──────────────────────────


def test_second_generation_uses_the_newer_json(session, tmp_path):
    """Deux générations successives dans la même session, JSON différents.

    La fusion est **gap-only** : si elle mutait le snapshot de session, les
    valeurs du premier JSON y resteraient et le second passage les verrait
    « déjà présentes ». Le second pack porterait alors des chiffres périmés
    tout en paraissant à jour.
    """
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    premier = tmp_path / "computed_1.json"
    _computed_json(premier)
    res1 = _generer(computed=str(premier), output_dir="pack_1")
    assert res1.get("status") != "error", res1
    assert res1["computed_quantities_coverage"]["n_merged"] == 7

    # Recalcul : mêmes éléments, valeurs différentes.
    second = tmp_path / "computed_2.json"
    doc = json.loads(premier.read_text(encoding="utf-8"))
    for q in doc["quantities"]:
        q["value"] = round(q["value"] * 2, 3)
    second.write_text(json.dumps(doc, ensure_ascii=False), encoding="utf-8")

    res2 = _generer(computed=str(second), output_dir="pack_2")
    assert res2.get("status") != "error", res2
    # Rien n'est « déjà présent » : la seconde fusion doit être complète.
    assert res2["computed_quantities_coverage"]["n_merged"] == 7
    assert res2["computed_quantities_coverage"]["n_gap_kept"] == 0

    shab2 = next(p for p in res2["paths"] if _cle(p) == "shab_xlsx")
    nombres = _nombres(shab2)
    assert any(abs(n - 49.0) < 0.01 for n in nombres), "valeur du 2e JSON attendue (24,5 x 2)"
    assert not any(abs(n - 24.5) < 0.01 for n in nombres), "valeur périmée du 1er JSON"


def test_session_snapshot_is_left_untouched(session, tmp_path):
    """La génération n'a pas d'effet de bord sur le snapshot de session."""
    sess, _ = session
    sess.snapshot = _snapshot_sans_quantites()

    _generer(computed=_computed_json(tmp_path / "computed.json"), output_dir="pack")

    espace = sess.snapshot.element_by_uuid["SP1"]
    assert not espace.get("computed_base_quantities")
    assert not espace.get("property_sets")

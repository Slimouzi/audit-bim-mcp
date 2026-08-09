"""Acceptation du pack AVP I3F — annexes non vides + formats MOA attendus.

Test d'acceptation déterministe (hors-ligne, CI) : sur un snapshot représentatif
(le chemin réel piloté par la maquette, ``sources=None``), le pack doit livrer les
annexes xlsx avec des lignes métier, dans le format MOA (pas de bandeau BIMData).
Aucune annexe ne doit porter l'ancienne charte (KORHUS). Double la garde runtime
``_qa_empty_deliverables`` d'une garantie de test permanente.
"""

from __future__ import annotations

import importlib.util
import zipfile
from pathlib import Path

import openpyxl
import pytest

from audit_bim.audit.engine import AuditResult
from audit_bim.audit.rules import audit_naming
from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.reporting.avp_i3f import (
    AvpQaError,
    _audit_stats,
    _count_business_rows,
    _count_controle_rows,
    write_avp_i3f_report_pack,
)
from audit_bim.reporting.bimdata_brand import WORDMARK
from audit_bim.reporting.theming import BIMDATA_FONT_PRIMARY, BIMDATA_PRIMARY
from audit_bim.reporting.word_report import NOT_AVAILABLE
from audit_bim.requirements.models import BIMPhase, NamingRule, RequirementsCatalog, ZoneSpec


def _catalog() -> RequirementsCatalog:
    return RequirementsCatalog(
        cch_version="3.6",
        cch_source_pdf="x",
        data_spec_source="x",
        naming_spec_source="x",
        properties=[],
        naming_rules=[],
        storey_names=[],
        zone_specs=[],
        room_specs=[],
    )


def _representative_result() -> AuditResult:
    """Maquette représentative peuplant **les 5 annexes** via le repli snapshot :
    espace (SHAB), zone (Zones/Espaces), mur d'enveloppe au calque réel
    ArchiCAD (Enveloppe), fenêtre + porte (Menuiseries) ; la Grille de contrôle
    (Contrôle) est toujours produite."""
    wall = {
        "uuid": "W1",
        "type": "IfcWall",
        "name": "Mur péri 221",
        "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetSideArea"}, "value": 30.0}],
            }
        ],
    }
    window = {
        "uuid": "WIN1",
        "type": "IfcWindow",
        "name": "F25",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [
                    {"definition": {"name": "Width"}, "value": 0.6},
                    {"definition": {"name": "Height"}, "value": 1.3},
                ],
            }
        ],
    }
    door = {"uuid": "D1", "type": "IfcDoor", "name": "P1"}
    slab = {
        "uuid": "SL1",
        "type": "IfcSlab",
        "name": "Dalle R+1",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetArea"}, "value": 12.98}],
            }
        ],
    }
    space = {
        "uuid": "S1",
        "type": "IfcSpace",
        "name": "CHAMBRE",
        "longname": "Chambre 01",
        "storey": {"uuid": "ST1", "name": "R+1"},
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 12.98}],
            }
        ],
    }
    zone = {"uuid": "Z1", "type": "IfcZone", "name": "Logement A101", "spaces": ["S1"]}
    snap = ModelSnapshot(
        project={"name": "Programme"},
        model={"name": "M.ifc"},
        storeys=[{"uuid": "ST1", "name": "R+1"}],
        spaces=[space],
        zones=[zone],
        elements=[wall, window, door, slab],
    ).index()
    return AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=snap, findings=[])


def _annexes(pack) -> dict[str, Path]:
    """Les annexes xlsx du pack, par libellé métier."""
    return {
        "Contrôle": pack.controle_xlsx,
        "SHAB": pack.shab_xlsx,
        "Zones/Espaces": pack.zones_espaces_xlsx,
        "Enveloppe": pack.enveloppe_xlsx,
        "Menuiseries": pack.menuiseries_xlsx,
        # Plancher n'est plus produit : bloqué par une règle métier absente.
        **({"Plancher": pack.plancher_xlsx} if pack.plancher_xlsx is not None else {}),
    }


def _xml_blob(path: Path) -> bytes:
    with zipfile.ZipFile(path) as z:
        return b"".join(z.read(n) for n in z.namelist() if n.endswith((".xml", ".rels"))).upper()


def _annex_rows(label: str, path: Path) -> int:
    """Compteur adapté : Contrôle a son compteur propre (lignes sous la grille),
    les 4 autres annexes utilisent le compteur générique de lignes métier."""
    return _count_controle_rows(path) if label == "Contrôle" else _count_business_rows(path)


def test_xlsx_annexes_are_non_empty(tmp_path):
    pack = write_avp_i3f_report_pack(
        _representative_result(),
        tmp_path / "out",
        sources=None,
        project_name="X",
        project_code="Y",
        export_pdf=False,
    )
    annexes = _annexes(pack)
    assert len(annexes) == 6
    empty = {label: p.name for label, p in annexes.items() if _annex_rows(label, p) == 0}
    assert not empty, f"annexes vides: {empty}"


def test_xlsx_annexes_use_moa_layout_not_bimdata_banner(tmp_path):
    pack = write_avp_i3f_report_pack(
        _representative_result(),
        tmp_path / "out",
        sources=None,
        project_name="X",
        project_code="Y",
        export_pdf=False,
    )
    for label, p in _annexes(pack).items():
        blob = _xml_blob(p)
        assert b"BIMDATA" not in blob, f"bandeau BIMDATA trouvé dans l'annexe MOA: {label}"
        assert b"KORHUS" not in blob, f"ancienne charte (KORHUS) trouvée: {label}"
        wb = openpyxl.load_workbook(p)
        assert not any(
            isinstance(c.value, str) and c.value.startswith("BIMDATA —")
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
        )
        wb.close()


def test_controle_grid_is_populated_from_audit(tmp_path):
    """Sans source I3F « Contrôle », la grille est générée depuis l'AuditResult
    (points de contrôle réels), donc le compteur propre est > 0."""
    pack = write_avp_i3f_report_pack(
        _representative_result(),
        tmp_path / "out",
        sources=None,
        project_name="X",
        project_code="Y",
        export_pdf=False,
    )
    assert _count_controle_rows(pack.controle_xlsx) > 0


def test_qa_gate_flags_truly_empty_controle(tmp_path):
    """Une grille de contrôle **réellement vide** (audit sans point de contrôle
    exploitable) déclenche ``AvpQaError`` — sans monkeypatch du compteur."""
    empty_snap = ModelSnapshot(project={"name": "P"}, model={"name": "M.ifc"}).index()
    result = AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=empty_snap, findings=[])
    with pytest.raises(AvpQaError) as exc:
        write_avp_i3f_report_pack(
            result,
            tmp_path / "out",
            sources=None,
            project_name="X",
            project_code="Y",
            export_pdf=False,
        )
    assert "Contrôle" in exc.value.empty


# ── Exactitude métier de la grille de contrôle (NO-GO #2) ─────────────────


def _catalog_zone_rules() -> RequirementsCatalog:
    """Catalogue avec pattern de Name de zone logement + liste d'ObjectType."""
    return RequirementsCatalog(
        cch_version="3.6",
        cch_source_pdf="x",
        data_spec_source="x",
        naming_spec_source="x",
        properties=[],
        naming_rules=[
            NamingRule(
                objet="Zone logement",
                ifc_class="IfcZone",
                ifc_attribute="Name",
                pattern=r"\d{4}L-\d{4}",
            )
        ],
        storey_names=[],
        zone_specs=[ZoneSpec(name="Logement T2", type_label="Zone Logement T2")],
        room_specs=[],
    )


def _zone_result(name, object_type) -> AuditResult:
    """AuditResult d'une seule IfcZone, findings issus de la VRAIE règle nommage."""
    cat = _catalog_zone_rules()
    zone = {"uuid": "Z1", "type": "IfcZone", "name": name, "object_type": object_type}
    snap = ModelSnapshot(project={"name": "P"}, model={"name": "M.ifc"}, zones=[zone]).index()
    findings = audit_naming(snap, cat)
    return AuditResult(phase=BIMPhase.AVP, catalog=cat, snapshot=snap, findings=findings)


def test_zone_bad_name_valid_objecttype_not_conflated():
    # Name invalide (zone logement, ne suit pas le pattern) mais ObjectType
    # valide → NON conforme en Nommage, CONFORME en ObjectType (pas d'agrégat).
    result = _zone_result("BADNAME", "Zone Logement T2")
    assert _audit_stats("Zones Nommage", result)["non_conforme"] == 1
    assert _audit_stats("Zones ObjectType", result)["non_conforme"] == 0


def test_zone_valid_name_absent_objecttype_not_conflated():
    # Name présent (ObjectType absent → pas de contrôle de pattern) mais
    # ObjectType manquant → CONFORME en Nommage, NON conforme en ObjectType.
    result = _zone_result("1802L-1101", None)
    assert _audit_stats("Zones Nommage", result)["non_conforme"] == 0
    assert _audit_stats("Zones ObjectType", result)["non_conforme"] == 1


def _material_result(material_list) -> AuditResult:
    wall = {"uuid": "W1", "type": "IfcWall", "name": "M"}
    if material_list is not None:
        wall["material_list"] = material_list
    snap = ModelSnapshot(project={"name": "P"}, model={"name": "M.ifc"}, elements=[wall]).index()
    return AuditResult(phase=BIMPhase.AVP, catalog=_catalog(), snapshot=snap, findings=[])


def test_material_present_is_conforme():
    # Mur avec matériau « Béton » (forme BIMData material_list) → conforme.
    result = _material_result([{"material": {"name": "Béton"}}])
    stats = _audit_stats("ARC absence de matériau", result)
    assert stats["non_conforme"] == 0
    assert stats["conforme"] == 1
    assert stats["conforme_ratio"] == 1.0


def test_material_absent_is_non_conforme():
    # Mur sans material_list → non conforme.
    result = _material_result(None)
    stats = _audit_stats("ARC absence de matériau", result)
    assert stats["non_conforme"] == 1
    assert stats["conforme"] == 0


def test_controle_grid_excel_cell_values(tmp_path):
    """Vérifie les **valeurs de cellules Excel** de la grille (pas que le nombre
    de lignes) : reproduit le cas CTO (Name invalide / ObjectType valide /
    matériau présent) et lit Total/Conformes/Non conformes dans le xlsx."""
    cat = _catalog_zone_rules()
    zone = {
        "uuid": "Z1",
        "type": "IfcZone",
        "name": "BADNAME",
        "object_type": "Zone Logement T2",
        "spaces": ["S1"],
    }
    space = {
        "uuid": "S1",
        "type": "IfcSpace",
        "name": "CHAMBRE",
        "longname": "Chambre",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 12.0}],
            }
        ],
    }
    wall = {
        "uuid": "W1",
        "type": "IfcWall",
        "name": "Mur",
        "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
        "material_list": [{"material": {"name": "Béton"}}],
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetSideArea"}, "value": 30.0}],
            }
        ],
    }
    snap = ModelSnapshot(
        project={"name": "P"},
        model={"name": "M.ifc"},
        spaces=[space],
        zones=[zone],
        elements=[wall],
    ).index()
    result = AuditResult(
        phase=BIMPhase.AVP, catalog=cat, snapshot=snap, findings=audit_naming(snap, cat)
    )
    pack = write_avp_i3f_report_pack(
        result, tmp_path / "out", sources=None, project_name="X", project_code="Y", export_pdf=False
    )
    wb = openpyxl.load_workbook(pack.controle_xlsx, data_only=True)
    ws = wb["Grille de contrôle"]
    # header MOA: [CODE 3F, POINTS..., EXIGENCE..., Outil..., EVALUATION, Commentaires]
    by_point = {
        row[1]: row
        for row in ws.iter_rows(values_only=True)
        if row
        and len(row) > 1
        and row[1] in ("Zones Nommage", "Zones ObjectType", "ARC bsence de matériau")
    }
    wb.close()
    assert by_point["Zones Nommage"][4] == 1  # Name invalide → évaluation à reprendre
    assert "non conformes: 1" in by_point["Zones Nommage"][5]
    assert by_point["Zones ObjectType"][4] == 2  # ObjectType valide (PAS confondu)
    assert "non conformes: 0" in by_point["Zones ObjectType"][5]
    assert by_point["ARC bsence de matériau"][4] == 2  # matériau Béton présent


# ── Acceptation du rapport Word (analyse BIM AVP) ─────────────────────────
#
# Un SEUL helper (``inspect_word_report`` du runner) porte les critères et les
# seuils, partagé entre le runner réseau et ces tests. Ici : cas POSITIF sur le
# vrai pack généré. Les cas NÉGATIFS (docx 1×1, projet absent, cellules
# NOT_AVAILABLE, section manquante) sont dans ``test_avp_acceptance_runner.py``.

_RUNNER_PATH = (
    Path(__file__).resolve().parents[2] / "scripts" / "avp_acceptance" / "run_acceptance.py"
)
_spec = importlib.util.spec_from_file_location("avp_acceptance_runner_pos", _RUNNER_PATH)
_runner = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_runner)


def test_word_report_accepted_on_real_pack(tmp_path):
    pack = write_avp_i3f_report_pack(
        _representative_result(),
        tmp_path / "out",
        sources=None,
        project_name="PROG-ACCEPT",
        project_code="0546L",
        phase="AVP",
        export_pdf=False,
    )
    result = _runner.inspect_word_report(
        pack.analyse_docx,
        expected_project_name="PROG-ACCEPT",
        phase="AVP",
        wordmark=WORDMARK,
        primary=BIMDATA_PRIMARY,
        font=BIMDATA_FONT_PRIMARY,
        not_available=NOT_AVAILABLE,
    )
    assert result["ok"] is True, result
    assert result["n_paragraphs"] >= 10
    assert result["n_significant_cells"] >= 10
    assert result["sections_ok"] is True
    assert result["moa_title"] is True
    assert result["annexes_present"] is True
    assert result["metadata_present"] is True
    assert result["no_korhus"] is True

"""Le catalogue AVP doit décrire le classeur que nous écrivons vraiment.

:mod:`avp_report_catalog` est purement déclaratif : il ne lit aucune donnée et
n'écrit aucun fichier. **Rien ne le contredit à l'exécution.** Un décalage y est
donc silencieux et durable — le livrable peut être corrigé pendant que le
contrat, lui, continue de promettre l'ancienne structure.

C'est arrivé deux fois. #210 a dû réaligner la spécification Menuiseries après
avoir restreint le livrable aux fenêtres. Puis le lot Zones/Espaces a ajouté un
onglet « Note de méthode », porté le détail à 12 colonnes et remplacé l'écart
absolu par un écart relatif gardé, sans que ``shab_maquette`` bouge : le
catalogue annonçait 2 onglets, 9 en-têtes et ``IF(Gn-Hn=0,"",Gn-Hn)``, une
formule que plus rien ne produisait.

Ce test ferme la boucle : il **génère** le pack et confronte les trois champs de
signature au classeur obtenu.
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pytest

from audit_bim.extraction.model_data import ModelSnapshot
from audit_bim.reporting.avp_i3f import write_avp_i3f_report_pack
from audit_bim.reporting.avp_report_catalog import REPORT_SPECS_BY_KEY

#: Rapports dont ``critical_formulas`` décrit encore la **géométrie du gabarit
#: MOA** (``SUM(D2:D10)``, ``COUNTA(D2:D16)``, ``E22/D22-1``…) et non notre
#: sortie. Dette explicite, héritée : la figer serait un défaut — le compteur
#: Fenêtres ne doit pas dépendre des 15 types du gabarit — mais la laisser
#: implicite en est un autre. Cet ensemble est asservi par un test : il ne peut
#: pas grandir sans faire échouer la suite.
_FORMULES_ENCORE_CELLES_DU_GABARIT = {"surface_enveloppe", "menuiseries"}

#: Rapports **bloqués par une règle métier** : ils n'écrivent aucun classeur,
#: donc il n'y a rien à confronter. Le catalogue continue de décrire la forme
#: qu'ils AURONT — supprimer cette description ferait perdre la cible.
_BLOQUES_SANS_CLASSEUR = {"plancher"}


def _snapshot() -> ModelSnapshot:
    """Maquette peuplant les cinq annexes générées depuis le snapshot.

    Reprend la forme de la maquette de recette : un espace zoné (SHAB,
    Zones/Espaces), un mur au calque d'enveloppe réel ArchiCAD (Enveloppe), une
    fenêtre (Fenêtres), une dalle (Plancher).
    """
    mur = {
        "uuid": "W1",
        "type": "IfcWall",
        "name": "Mur péri 221",
        "layers": [{"name": "221 - MURS - Extérieurs périphériques.Exndo"}],
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetSideArea"}, "value": 30.0}],
            }
        ],
    }
    fenetre = {
        "uuid": "WIN1",
        "type": "IfcWindow",
        "name": "F25",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [
                    {"definition": {"name": "Width"}, "value": 0.6},
                    {"definition": {"name": "Height"}, "value": 1.3},
                ],
            }
        ],
    }
    dalle = {
        "uuid": "SL1",
        "type": "IfcSlab",
        "name": "Dalle R+1",
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetArea"}, "value": 12.98}],
            }
        ],
    }
    espace = {
        "uuid": "S1",
        "type": "IfcSpace",
        "name": "CHAMBRE",
        "longname": "Chambre 01",
        "storey": {"uuid": "ST1", "name": "R+1"},
        "property_sets": [
            {
                "name": "BaseQuantities",
                "properties": [{"definition": {"name": "NetFloorArea"}, "value": 12.98}],
            }
        ],
    }
    zone = {
        "uuid": "Z1",
        "type": "IfcZone",
        "name": "Logement A101",
        "spaces": ["S1"],
        "attributes": {
            "properties": [{"definition": {"name": "ObjectType"}, "value": "Zone Logement T2"}]
        },
    }
    return ModelSnapshot(
        project={"name": "Programme"},
        model={"name": "M.ifc"},
        storeys=[{"uuid": "ST1", "name": "R+1"}],
        spaces=[espace],
        zones=[zone],
        elements=[mur, fenetre, dalle, zone],
    ).index()


@pytest.fixture(scope="module")
def classeurs(tmp_path_factory) -> dict[str, Path]:
    """Les cinq annexes générées depuis le snapshot, par clé de catalogue."""
    pack = write_avp_i3f_report_pack(
        None,
        tmp_path_factory.mktemp("pack"),
        snapshot=_snapshot(),
        project_name="P",
        project_code="C",
        export_pdf=False,
    )
    return {
        "shab_maquette": pack.shab_xlsx,
        "zones_espaces": pack.zones_espaces_xlsx,
        "surface_enveloppe": pack.enveloppe_xlsx,
        "menuiseries": pack.menuiseries_xlsx,
    }


_CLES_TOUTES = ("shab_maquette", "zones_espaces", "surface_enveloppe", "menuiseries", "plancher")
_CLES = tuple(c for c in _CLES_TOUTES if c not in _BLOQUES_SANS_CLASSEUR)


def _normalise(formule: str) -> str:
    """``=IF(H2/G2-1…)`` → ``IF(Hn/Gn-1…)`` — comparable ligne à ligne."""
    return re.sub(r"([A-Z]{1,2})\d+", r"\1n", str(formule)).lstrip("=")


def _onglets(chemin: Path) -> tuple[str, ...]:
    return tuple(openpyxl.load_workbook(chemin).sheetnames)


def _entetes(chemin: Path) -> list[str]:
    wb = openpyxl.load_workbook(chemin)
    return [str(c.value) for t in wb.sheetnames for c in wb[t][1] if c.value is not None]


def _formules(chemin: Path) -> set[str]:
    wb = openpyxl.load_workbook(chemin, data_only=False)
    return {
        _normalise(c.value)
        for t in wb.sheetnames
        for row in wb[t].iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    }


@pytest.mark.parametrize("cle", _CLES)
def test_les_onglets_annonces_sont_ceux_qui_sortent(classeurs, cle):
    """Égalité, pas inclusion : c'est l'ajout d'un onglet — « Note de méthode »
    — qui a rendu la spécification SHAB fausse, et une inclusion l'aurait laissé
    passer."""
    spec = REPORT_SPECS_BY_KEY[cle]
    produits = _onglets(classeurs[cle])
    assert tuple(spec.expected_sheets) == produits, (
        f"{cle} : le catalogue annonce {spec.expected_sheets}, le classeur porte {produits}"
    )


@pytest.mark.parametrize("cle", _CLES)
def test_aucun_entete_annonce_nest_absent_du_classeur(classeurs, cle):
    spec = REPORT_SPECS_BY_KEY[cle]
    reels = _entetes(classeurs[cle])
    absents = [h for h in spec.headers if h not in reels]
    assert not absents, f"{cle} : en-têtes promis mais absents du classeur — {absents}"


@pytest.mark.parametrize("cle", sorted(set(_CLES) - _FORMULES_ENCORE_CELLES_DU_GABARIT))
def test_les_formules_critiques_annoncees_sont_ecrites(classeurs, cle):
    """Comparaison **exacte**, pas par sous-chaîne.

    L'écart relatif du gabarit ``IF(Hn/Gn-1=0,"",Hn/Gn-1)`` est une sous-chaîne
    de la formule gardée que nous écrivons : une comparaison laxiste aurait
    déclaré la spécification à jour alors qu'elle décrivait une formule que nous
    n'écrivons plus telle quelle.
    """
    spec = REPORT_SPECS_BY_KEY[cle]
    produites = _formules(classeurs[cle])
    assert spec.critical_formulas, f"{cle} : aucune formule critique déclarée"
    for formule in spec.critical_formulas:
        assert formule in produites, (
            f"{cle} : formule annoncée « {formule} » jamais écrite. "
            f"Le classeur porte {sorted(produites)}"
        )


def test_la_dette_des_formules_de_gabarit_ne_grandit_pas(classeurs):
    """Non-vacuité de l'exception, et verrou contre son élargissement.

    Trois rapports décrivent encore la géométrie du gabarit MOA dans un champ
    qui, partout ailleurs, décrit notre sortie. La dette est nommée plutôt que
    masquée : si un quatrième rapport y tombe — ou si l'un des trois est
    réaligné — ce test le dit.
    """
    decalees = {
        cle
        for cle in _CLES
        if any(
            f not in _formules(classeurs[cle]) for f in REPORT_SPECS_BY_KEY[cle].critical_formulas
        )
    }
    assert decalees == _FORMULES_ENCORE_CELLES_DU_GABARIT, (
        f"périmètre de la dette modifié : {decalees} au lieu de "
        f"{_FORMULES_ENCORE_CELLES_DU_GABARIT}"
    )


@pytest.mark.parametrize("cle", ("shab_maquette", "zones_espaces"))
def test_le_detail_zones_espaces_est_annonce_en_entier(classeurs, cle):
    """Les 12 colonnes A:L sont la forme du gabarit client : les annoncer
    partiellement, c'est promettre un autre tableau. Une inclusion ne suffit pas
    ici — c'est l'omission de ``Groupes`` / ``Couleur`` / ``écarts`` qui rendait
    la spécification SHAB fausse.

    La comparaison porte sur la **ligne d'en-tête écrite dans le classeur**, et
    non sur la constante du builder : confronter une déclaration à une autre
    déclaration ne prouve que leur accord, pas leur vérité.
    """
    wb = openpyxl.load_workbook(classeurs[cle])
    detail = next(t for t in wb.sheetnames if t.startswith("TDB 2022 01.3"))
    ecrits = [c.value for c in wb[detail][1]]
    assert list(REPORT_SPECS_BY_KEY[cle].headers) == ecrits, (
        f"{cle} : le catalogue annonce {len(REPORT_SPECS_BY_KEY[cle].headers)} colonnes, "
        f"le classeur en écrit {len(ecrits)}"
    )


def test_un_rapport_bloque_ecrit_bien_aucun_classeur(classeurs):
    """Non-vacuité de l'exclusion : si `plancher` produisait de nouveau un
    fichier, ce test le dirait — et le garde catalogue↔sortie devrait le
    reprendre en charge."""
    from audit_bim.reporting.avp_report_catalog import REPORT_SPECS_BY_KEY

    for cle in _BLOQUES_SANS_CLASSEUR:
        assert cle not in classeurs
        assert REPORT_SPECS_BY_KEY[cle].blocked_reason, (
            f"{cle} est exclu du garde sans être déclaré bloqué"
        )
